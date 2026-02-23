"""
EY ServiceEdge — Flask API Server (Enhanced)
CR-019: Full downstream reactivity on all mutation endpoints.
All engines integrated with proper data mapping.
"""
import json
import os
import traceback
from flask import Flask, jsonify, request, render_template, send_file
from engines.data_loader import run_etl
from engines.diagnostic import run_diagnostic
from engines.maturity import run_maturity
from engines.readiness import compute_readiness, STRATEGIC_DRIVERS
from engines.waterfall import score_initiatives, run_waterfall
from engines.risk import run_risk
from engines.workforce import run_workforce
from engines.channel_strategy import run_channel_strategy

app = Flask(__name__)

STATE = {
    'data': None, 'diagnostic': None, 'maturity': None,
    'readiness': None, 'initiatives': None, 'waterfall': None,
    'risk': None, 'workforce': None, 'channelStrategy': None,
    'overrides': {}, 'loaded': False,
}


def _sanitize_for_json(obj):
    """Convert sets and other non-JSON-serializable types to lists."""
    if isinstance(obj, set):
        return sorted(list(obj))
    elif isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    return obj


def _run_all():
    """Run full pipeline with proper downstream chaining."""
    data = run_etl()
    if 'totalCost' not in data:
        data['totalCost'] = sum(r['headcount'] * r['costPerFTE'] for r in data['roles'])
    if 'avgCPC' not in data:
        # CR-021 v5: CPC = annual cost / annual volume (not raw volume)
        annual_vol = data.get('totalVolumeAnnual', data.get('totalVolume', 1))
        data['avgCPC'] = round(data['totalCost'] / max(annual_vol, 1), 2)
    STATE['data'] = data

    diagnostic = run_diagnostic(data)
    STATE['diagnostic'] = diagnostic

    maturity = run_maturity(data, diagnostic)
    STATE['maturity'] = maturity

    readiness_ctx = compute_readiness(data, diagnostic)
    STATE['readiness'] = readiness_ctx

    initiatives = score_initiatives(data, diagnostic, readiness_ctx)
    STATE['initiatives'] = initiatives
    _apply_all_overrides()

    waterfall = run_waterfall(data, initiatives)
    STATE['waterfall'] = waterfall

    risk = run_risk(initiatives, data)
    STATE['risk'] = risk

    workforce = run_workforce(data, waterfall, initiatives)
    STATE['workforce'] = workforce

    channel_strategy = run_channel_strategy(data, diagnostic)
    STATE['channelStrategy'] = channel_strategy

    STATE['loaded'] = True
    return True


def _recompute_downstream():
    """CR-019: Recompute waterfall→risk→workforce after initiative changes."""
    data = STATE['data']
    STATE['waterfall'] = run_waterfall(data, STATE['initiatives'])
    STATE['risk'] = run_risk(STATE['initiatives'], data)
    STATE['workforce'] = run_workforce(data, STATE['waterfall'], STATE['initiatives'])


def _recompute_all_from_diagnostic():
    """Full recompute from diagnostic onwards, using EY readiness engine.
    CR-019b: Preserves ALL user overrides (impact, adoption, channels, roles, etc.)"""
    data = STATE['data']
    STATE['diagnostic'] = run_diagnostic(data)
    STATE['maturity'] = run_maturity(data, STATE['diagnostic'])

    readiness_ctx = compute_readiness(data, STATE['diagnostic'])
    STATE['readiness'] = readiness_ctx

    STATE['initiatives'] = score_initiatives(data, STATE['diagnostic'], readiness_ctx)
    _apply_all_overrides()
    _recompute_downstream()
    STATE['channelStrategy'] = run_channel_strategy(data, STATE['diagnostic'])


def _apply_all_overrides():
    """Apply ALL stored overrides to initiatives — enabled, ramp, and field edits."""
    for init in STATE['initiatives']:
        iid = init['id']
        # Enabled state
        ok = f"init_enabled_{iid}"
        if ok in STATE['overrides']:
            init['enabled'] = STATE['overrides'][ok]
        # Ramp overrides
        for rk in ('rampYear1','rampYear2','rampYear3'):
            rkey = f"init_{rk}_{iid}"
            if rkey in STATE['overrides']:
                init[rk] = STATE['overrides'][rkey]
        # Field overrides (impact, adoption, channels, roles, lever, effort, etc.)
        field_key = f"init_fields_{iid}"
        if field_key in STATE['overrides']:
            fields = STATE['overrides'][field_key]
            for fk, fv in fields.items():
                init[fk] = fv


@app.before_request
def _ensure_loaded():
    if not STATE['loaded'] and not STATE.get('_load_error'):
        try:
            _run_all()
            print("[OK] ServiceEdge engines loaded successfully")
        except Exception as e:
            err_msg = f"{type(e).__name__}: {e}"
            STATE['_load_error'] = err_msg
            print(f"\n{'='*60}")
            print(f"[!] ENGINE LOAD FAILED — Frontend will show demo data!")
            print(f"[!] Error: {err_msg}")
            print(f"[!] Fix the issue above, then restart python app.py")
            print(f"{'='*60}\n")
            traceback.print_exc()


def _build_demo_object(overrides=None):
    """Build the DEMO object the frontend expects — maps all engine outputs.
    
    Args:
        overrides: Optional dict with keys like 'waterfall', 'initiatives', 'risk', 'workforce'
                   to override STATE values (used for scoped layer calculations).
    """
    ov = overrides or {}
    data = STATE['data']; diag = STATE['diagnostic']; mat = STATE['maturity']
    wf = ov.get('waterfall', STATE['waterfall'])
    rsk = ov.get('risk', STATE['risk'])
    wkf = ov.get('workforce', STATE['workforce'])
    chs = STATE['channelStrategy']; roles = data['roles']
    inits = ov.get('initiatives', STATE['initiatives'])

    return {
        # Core data
        'queues': data['queues'], 'roles': roles, 'params': data['params'],
        'benchmarks': data.get('benchmarks', {}),
        'totalVolume': data['totalVolume'], 'totalFTE': data['totalFTE'],
        'totalCost': data['totalCost'],
        'totalVolumeAnnual': data.get('totalVolumeAnnual', data['totalVolume']),
        'volumeAnnualizationFactor': data.get('volumeAnnualizationFactor', 12),
        'totalMonthlyCost': sum(r['headcount']*r['costPerFTE']/12 for r in roles),
        'avgCSAT': data['avgCSAT'], 'avgAHT': round(data['avgAHT'] * 60, 0),  # seconds for UI
        'avgAHT_min': round(data['avgAHT'], 1),  # minutes for reference
        'avgFCR': data.get('avgFCR', 0), 'avgCPC': data.get('avgCPC', 0),
        # Breakdowns
        'channelMix': _build_channel_mix(data['queues']),
        'buMix': _build_bu_mix(data['queues']),
        'intentMix': _build_intent_mix(data['queues']),
        # Diagnostic
        'healthScores': diag.get('queueScores', []),
        'overallHealth': diag.get('summary', {}).get('avgScore', 0),
        'healthRag': 'green' if diag.get('summary',{}).get('avgScore',0) >= 70 else 'amber' if diag.get('summary',{}).get('avgScore',0) >= 40 else 'red',
        'problemAreas': diag.get('problemAreas', []),
        'rootCauses': diag.get('rootCauses', []),
        'costAnalysis': diag.get('costAnalysis', {}),
        'channelSummary': diag.get('channelSummary', []),
        'mismatch': diag.get('mismatch', []),
        'mismatchSummary': diag.get('mismatch', []),
        'mismatchDetail': diag.get('mismatch', []),
        # Maturity
        'maturityScores': mat.get('dimensions', {}),
        'maturityOverall': mat.get('overall', 0),
        'maturityLevel': mat.get('overallLevel', 1),
        'maturityLabel': mat.get('levelInfo', {}).get('label', ''),
        'maturityGaps': mat.get('gaps', []),
        'maturityRadar': mat.get('radar', {}),
        'maturityRecommendations': mat.get('gaps', []),
        # Initiatives
        'initiatives': inits,
        'enabledCount': sum(1 for i in inits if i.get('enabled')),
        # Waterfall
        'waterfall': wf,
        'yearlyProjections': wf.get('yearly', []),
        'scenarios': wf.get('scenarios', {}),
        'sensitivity': wf.get('sensitivity', []),
        'roleImpact': wf.get('roleImpact', {}),
        'investmentItems': wf.get('investmentItems', []),
        'investmentYearly': wf.get('investmentYearly', []),
        'investmentSummary': wf.get('investmentSummary', {}),
        'financials': {
            'totalNPV': wf.get('totalNPV',0), 'totalSaving': wf.get('totalSaving',0),
            'totalInvestment': wf.get('totalInvestment',0), 'roi': wf.get('roi',0),
            'roiGross': wf.get('roiGross',0), 'payback': wf.get('payback',0),
            'irr': wf.get('irr',0), 'techInvestment': wf.get('techInvestment',0),
            'annualMaintenance': wf.get('annualMaintenance',0),
        },
        # Risk
        'riskRegister': rsk.get('initiatives', []),
        'riskSummary': rsk.get('summary', {}),
        'topRisks': sorted(rsk.get('initiatives',[]), key=lambda x: x.get('overallRisk',0), reverse=True)[:5],
        # Workforce
        'workforceTransition': wkf.get('transitions', []),
        'workforceSummary': wkf.get('summary', {}),
        'reskillMatrix': wkf.get('reskillMatrix', {}),
        # Heatmap
        'heatmapData': _build_heatmap(data['queues']),
        'costBreakdown': _build_cost_breakdown(data),
        # Channel Strategy
        'channelRecommendations': chs.get('recommendations', []),
        'channelSankey': chs.get('sankey', {}),
        'currentDigitalPct': chs.get('currentDigitalPct', 0),
        'targetDigitalPct': chs.get('targetDigitalPct', 0),
        'channelScorecard': chs.get('recommendations', []),
        'channelIntroductions': [r for r in chs.get('recommendations',[]) if r.get('decision')=='invest'],
        'channelRetirements': [r for r in chs.get('recommendations',[]) if r.get('decision') in ('sunset','migrate_from')],
        'channelOptimalMix': {'currentDigital': chs.get('currentDigitalPct',0), 'targetDigital': chs.get('targetDigitalPct',0)},
        'intentMatrix': chs.get('intentMatrix', []),
        'targetMix': chs.get('targetMix', {}),
        'channelMigrations': chs.get('migrations', []),
        'migrationReadiness': chs.get('migrationReadiness', {}),
        'frictionSignals': chs.get('frictionSignals', []),
        'cxSafeguards': chs.get('cxSafeguards', {}),
        'channelCostAnalysis': chs.get('costAnalysis', {}),
        'migrationSavings': chs.get('migrationSavings', {}),
        # Readiness Scores (EY Methodology)
        'readiness': _sanitize_for_json(STATE.get('readiness', {})),
        'automationReadiness': STATE.get('readiness', {}).get('automationReadiness', 0),
        'opModelGap': STATE.get('readiness', {}).get('opModelGap', 0),
        'locationScore': STATE.get('readiness', {}).get('locationScore', 0),
        'strategicDriver': STATE.get('readiness', {}).get('strategicDriver', 'cost_optimization'),
        'strategicDrivers': STRATEGIC_DRIVERS,
        # Layer FTE breakdown (for FTE waterfall chart)
        'layerFTE': wf.get('layerFTE', {}),
        'layerSaving': wf.get('layerSaving', {}),
        # Pool-based engine v4
        'poolUtilization': wf.get('poolUtilization', {}),
        'poolSummary': wf.get('poolSummary', {}),
        'auditTrail': wf.get('auditTrail', []),
        # Meta
        'clientName': data['params'].get('clientName','Client'),
        'industry': data['params'].get('industry','Custom'),
        'currency': data['params'].get('currency','USD'),
        'horizon': data['params'].get('horizon',3),
    }


# ══════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/')
def index():
    if STATE['loaded']:
        server_data = {
            'demo': _build_demo_object(),
            'initiatives': STATE['initiatives'],
            'waterfall': STATE['waterfall'],
        }
        return render_template('index.html', server_data=json.dumps(server_data, default=str),
                               load_error=None)
    return render_template('index.html', server_data=None,
                           load_error=STATE.get('_load_error', None))


@app.route('/api/data')
def api_data():
    if not STATE['loaded']:
        return jsonify({
            'error': 'Data not loaded',
            'reason': STATE.get('_load_error', 'Unknown — check terminal'),
            'hint': 'Common fixes: (1) Move folder out of OneDrive, (2) cd into project folder before running, (3) pip install openpyxl'
        }), 503
    return jsonify(_build_demo_object())


@app.route('/api/diagnostic')
def api_diagnostic():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['diagnostic'])

@app.route('/api/maturity')
def api_maturity():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['maturity'])

@app.route('/api/channel-strategy')
def api_channel_strategy():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['channelStrategy'])

@app.route('/api/initiatives')
def api_initiatives():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify({
        'initiatives': STATE['initiatives'],
        'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
        'totalCount': len(STATE['initiatives']),
    })

@app.route('/api/waterfall')
def api_waterfall():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['waterfall'])

@app.route('/api/risk')
def api_risk():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['risk'])

@app.route('/api/workforce')
def api_workforce():
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    return jsonify(STATE['workforce'])


@app.route('/api/refresh', methods=['POST'])
def api_refresh():
    """Full reload from Excel — resets ALL overrides, complete fresh start."""
    try:
        STATE['overrides'] = {}
        STATE['loaded'] = False
        STATE['_load_error'] = None  # Clear previous error so retry works
        _run_all()
        return jsonify({'status':'ok','message':'All engines refreshed from source data',
                        'data': _build_demo_object(),
                        'initiatives': STATE['initiatives'],
                        'waterfall': STATE['waterfall']})
    except Exception as e:
        return jsonify({'status':'error','message':str(e)}), 500


@app.route('/api/recalculate', methods=['POST'])
def api_recalculate():
    """Re-run computation pipeline preserving current overrides/edits.
    Accepts optional params to update, optional activeLayer for scoped calc.
    Change 5b: frontend sends full parameter state here.
    Change 6b: activeLayer scopes calculation to one layer only.
    Change 9: distinct from /api/refresh which resets everything.
    """
    try:
        body = request.get_json(force=True) if request.is_json else {}
        data = STATE['data']

        # ── Apply any parameter updates from frontend ──
        params_update = body.get('params', {})
        for key, value in params_update.items():
            if key in data['params']:
                data['params'][key] = value
                STATE['overrides'][key] = value

        # ── Strategic driver update ──
        if 'strategicDriver' in params_update:
            data['params']['strategicDriver'] = params_update['strategicDriver']

        # ── Full recompute from diagnostic ──
        _recompute_all_from_diagnostic()

        # ── Optional: scoped layer calculation (Change 6b) ──
        active_layer = body.get('activeLayer')
        if active_layer and active_layer != 'All Layers':
            # Temporarily disable initiatives outside the active layer
            import copy
            scoped_inits = copy.deepcopy(STATE['initiatives'])
            for init in scoped_inits:
                if init.get('layer') != active_layer:
                    init['enabled'] = False
            scoped_waterfall = run_waterfall(data, scoped_inits)
            scoped_risk = run_risk(scoped_inits, data)
            scoped_workforce = run_workforce(data, scoped_waterfall, scoped_inits)

            return jsonify({
                'status': 'ok',
                'scoped': True,
                'activeLayer': active_layer,
                'data': _build_demo_object(overrides={
                    'waterfall': scoped_waterfall,
                    'risk': scoped_risk,
                    'workforce': scoped_workforce,
                    'initiatives': scoped_inits,
                }),
                'initiatives': scoped_inits,
                'waterfall': scoped_waterfall,
                'risk': scoped_risk,
                'workforce': scoped_workforce,
                'enabledCount': sum(1 for i in scoped_inits if i.get('enabled')),
            })

        return jsonify({
            'status': 'ok',
            'scoped': False,
            'data': _build_demo_object(),
            'initiatives': STATE['initiatives'],
            'waterfall': STATE['waterfall'],
            'risk': STATE['risk'],
            'workforce': STATE['workforce'],
            'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'status':'error','message':str(e)}), 500


@app.route('/api/initiative/toggle', methods=['POST'])
def api_toggle_initiative():
    """CR-019: Enable/disable initiative + full downstream recompute."""
    body = request.get_json(force=True)
    init_id = body.get('id'); enabled = body.get('enabled')
    if not init_id or enabled is None:
        return jsonify({'error':'id and enabled required'}), 400

    STATE['overrides'][f"init_enabled_{init_id}"] = bool(enabled)
    for init in STATE['initiatives']:
        if init['id'] == init_id:
            init['enabled'] = bool(enabled); break

    _recompute_downstream()
    return jsonify({
        'status':'ok',
        'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
        'waterfall': STATE['waterfall'], 'risk': STATE['risk'],
        'workforce': STATE['workforce'], 'initiatives': STATE['initiatives'],
    })


@app.route('/api/initiative/update', methods=['POST'])
def api_update_initiative():
    """CR-019b: Persist initiative field edits (impact, adoption, channels, etc.)
    across recalculations. Stores as field-level overrides."""
    body = request.get_json(force=True)
    init_id = body.get('id')
    fields = body.get('fields', {})
    if not init_id:
        return jsonify({'error': 'id required'}), 400

    # Store field overrides persistently
    field_key = f"init_fields_{init_id}"
    if field_key not in STATE['overrides']:
        STATE['overrides'][field_key] = {}
    STATE['overrides'][field_key].update(fields)

    # Apply to current initiative immediately
    for init in STATE['initiatives']:
        if init['id'] == init_id:
            for fk, fv in fields.items():
                init[fk] = fv
            break

    # Recompute downstream
    _recompute_downstream()
    return jsonify({
        'status': 'ok',
        'message': f'Initiative {init_id} updated with {len(fields)} field(s)',
        'enabledCount': sum(1 for i in STATE['initiatives'] if i.get('enabled')),
        'waterfall': STATE['waterfall'], 'risk': STATE['risk'],
        'workforce': STATE['workforce'], 'initiatives': STATE['initiatives'],
    })


@app.route('/api/override', methods=['POST'])
def api_override():
    """CR-019: Parameter override + full recompute."""
    body = request.get_json(force=True)
    key = body.get('key'); value = body.get('value')
    if not key: return jsonify({'error':'key required'}), 400

    STATE['overrides'][key] = value
    if key in STATE['data']['params']:
        STATE['data']['params'][key] = value

    _recompute_all_from_diagnostic()
    return jsonify({'status':'ok','message':f'Override {key}={value} applied',
                    'data': _build_demo_object(),
                    'initiatives': STATE['initiatives'],
                    'waterfall': STATE['waterfall']})


@app.route('/api/maturity/override', methods=['POST'])
def api_maturity_override():
    """Override maturity dimension score."""
    body = request.get_json(force=True)
    dimension = body.get('dimension'); score = body.get('score')
    if not dimension or score is None:
        return jsonify({'error':'dimension and score required'}), 400

    score = max(1.0, min(5.0, float(score)))
    STATE['overrides'][f"maturity_{dimension}"] = score

    mat = STATE['maturity']
    dims = mat.get('dimensions', {})
    if dimension in dims:
        dims[dimension]['score'] = score
        dims[dimension]['level'] = min(5, max(1, round(score)))
    # Recalc overall
    if dims:
        from engines.maturity import DIMENSIONS, MATURITY_LEVELS
        overall = sum(dims[k]['score'] * dims[k].get('weight', 0.20) for k in dims if isinstance(dims[k], dict))
        mat['overall'] = round(overall, 2)
        mat['overallLevel'] = min(5, max(1, round(overall)))
        mat['levelInfo'] = MATURITY_LEVELS.get(mat['overallLevel'], {})

    return jsonify({'status':'ok','maturity': mat})


@app.route('/api/investment')
def api_investment():
    """CR-017: Investment breakdown from waterfall engine."""
    if not STATE['loaded']: return jsonify({'error':'Not loaded'}), 503
    wf = STATE['waterfall']
    return jsonify({
        'items': wf.get('investmentItems', []),
        'summary': wf.get('investmentSummary', {}),
        'yearly': wf.get('investmentYearly', []),
    })


@app.route('/api/initiatives/batch', methods=['POST'])
def api_batch_initiatives():
    """CR-013: Batch update with ramp-up controls."""
    body = request.get_json(force=True)
    updates = body.get('updates', [])
    for upd in updates:
        iid = upd.get('id')
        for init in STATE['initiatives']:
            if init['id'] == iid:
                if 'enabled' in upd:
                    init['enabled'] = bool(upd['enabled'])
                    STATE['overrides'][f"init_enabled_{iid}"] = bool(upd['enabled'])
                for rk in ('rampYear1','rampYear2','rampYear3'):
                    if rk in upd:
                        init[rk] = float(upd[rk])
                        STATE['overrides'][f"init_{rk}_{iid}"] = float(upd[rk])
                if 'priority' in upd:
                    init['priority'] = upd['priority']
                break

    _recompute_downstream()
    return jsonify({'status':'ok','initiatives': STATE['initiatives'],'waterfall': STATE['waterfall']})


@app.route('/api/export')
def api_export():
    """Export all data to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        hf = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='2E2E38', end_color='2E2E38', fill_type='solid')
        tb = Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))

        def ws_write(ws, headers, rows):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font=hf; cell.fill=hfill; cell.alignment=Alignment(horizontal='center'); cell.border=tb
            for r, row in enumerate(rows, 2):
                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c, value=val); cell.border=tb
            for col in ws.columns:
                ml = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(ml+2, 40)

        data=STATE['data']; diag=STATE['diagnostic']; wf=STATE['waterfall']

        # 1. Executive Summary
        ws=wb.active; ws.title='Executive Summary'
        ws_write(ws, ['Metric','Value'], [
            ['Client', data['params'].get('clientName','')],
            ['Industry', data['params'].get('industry','')],
            ['Total Volume (raw)', data['totalVolume']],
            ['Volume Ann. Factor', data.get('volumeAnnualizationFactor', 12)],
            ['Total Volume (annual)', data.get('totalVolumeAnnual', data['totalVolume'])],
            ['Total FTE', data['totalFTE']],
            ['Total Cost', f"${data['totalCost']:,.0f}"],
            ['Avg CSAT', f"{data['avgCSAT']:.2f}"],
            ['Avg AHT', f"{data['avgAHT']:.1f} min ({data['avgAHT'] * 60:.0f}s)"],
            ['Avg FCR', f"{data.get('avgFCR',0):.1%}"],
            ['Avg CPC', f"${data.get('avgCPC',0):.2f}"],
            ['NPV', f"${wf.get('totalNPV',0):,.0f}"],
            ['IRR', f"{wf.get('irr',0):.1f}%"],
            ['Total Investment', f"${wf.get('totalInvestment',0):,.0f}"],
        ])

        # 2. Queue Detail
        ws2=wb.create_sheet('Queue Detail')
        ws_write(ws2, ['Queue','Channel','Volume','CSAT','FCR','AHT','CPC','Escalation'], [
            [q['queue'],q['channel'],q['volume'],f"{q.get('csat',0):.2f}",f"{q.get('fcr',0):.1%}",
             f"{q.get('aht',0):.0f}",f"${q.get('cpc',0):.2f}",f"{q.get('escalation',0):.1%}"]
            for q in data['queues']
        ])

        # 3. Health Scores
        ws3=wb.create_sheet('Health Scores')
        ws_write(ws3, ['Queue','Channel','Volume','Score','Rating'], [
            [h['queue'],h['channel'],h['volume'],f"{h['overallScore']:.1f}",h['rating']]
            for h in diag.get('queueScores', [])
        ])

        # 4. Root Causes
        ws4=wb.create_sheet('Root Causes')
        ws_write(ws4, ['Queue','Channel','Rating','Worst Metric','Score','Root Cause','Recommendation'], [
            [rc['queue'],rc['channel'],rc['rating'],rc['worstMetric'],f"{rc['score']:.1f}",
             rc['rootCause'],rc['recommendation']]
            for rc in diag.get('rootCauses', [])
        ])

        # 5. Initiatives
        ws5=wb.create_sheet('Initiatives')
        ws_write(ws5, ['ID','Name','Layer','Lever','Enabled','Score','FTE Impact','Annual Saving','Contribution%'], [
            [i['id'],i['name'],i['layer'],i['lever'],'Yes' if i.get('enabled') else 'No',
             f"{i.get('matchScore',0):.1f}",f"{i.get('_fteImpact',0):.1f}",
             f"${i.get('_annualSaving',0):,.0f}",f"{i.get('_contributionPct',0):.1f}%"]
            for i in STATE['initiatives']
        ])

        # 6. Waterfall
        ws6=wb.create_sheet('Waterfall')
        ws_write(ws6, ['Year','FTE Reduction','Final FTE','Annual Saving','Cum Saving','NPV'], [
            [y['year'],y['fteReduction'],y['finalFTE'],f"${y['annualSaving']:,.0f}",
             f"${y['cumSaving']:,.0f}",f"${y['npv']:,.0f}"]
            for y in wf.get('yearly', [])
        ])

        # 7. Investment
        ws7=wb.create_sheet('Investment')
        ws_write(ws7, ['ID','Name','Layer','One-Time','Recurring','Source'], [
            [it['id'],it['name'],it['layer'],f"${it['oneTime']:,.0f}",f"${it['recurring']:,.0f}",it['source']]
            for it in wf.get('investmentItems', [])
        ])

        # 8. Risk
        ws8=wb.create_sheet('Risk')
        ws_write(ws8, ['ID','Name','Layer','Risk Score','Rating','Mitigations'], [
            [r['id'],r['name'],r['layer'],f"{r['overallRisk']:.2f}",r['rating'],
             '; '.join(r.get('mitigations',[]))]
            for r in STATE['risk'].get('initiatives', [])
        ])

        # 9. Workforce
        ws9=wb.create_sheet('Workforce')
        ws_write(ws9, ['Role','Year','Baseline','Reduction','Attrited','Redeployed','Separated','Transition Cost'], [
            [t['role'],t['year'],t['baseline'],t['reduction'],t['attrited'],t['redeployed'],
             t['separated'],f"${t['totalTransitionCost']:,.0f}"]
            for t in STATE['workforce'].get('transitions', [])
        ])

        # 10. Maturity
        ws10=wb.create_sheet('Maturity')
        mat = STATE['maturity']
        dims = mat.get('dimensions', {})
        ws_write(ws10, ['Dimension','Score','Level','Description'], [
            [d.get('label',''),f"{d.get('score',0):.2f}",d.get('level',0),d.get('description','')]
            for d in (dims.values() if isinstance(dims, dict) else dims)
        ])

        # 11. Channel Strategy
        ws11=wb.create_sheet('Channel Strategy')
        ws_write(ws11, ['Channel','Decision','Vol Share','Avg Score','Avg CPC','Rationale'], [
            [r['channel'],r['decision'],f"{r['volumeShare']:.1f}%",f"{r['avgScore']:.1f}",
             f"${r['avgCpc']:.2f}",r['rationale'][:100]]
            for r in STATE['channelStrategy'].get('recommendations', [])
        ])

        # 12. Scenarios
        ws12=wb.create_sheet('Scenarios')
        ws_write(ws12, ['Scenario','NPV','FTE Reduction','Investment','IRR','Annual Saving'], [
            [s.get('label',''),f"${s.get('npv',0):,.0f}",s.get('fteReduction',0),
             f"${s.get('investment',0):,.0f}",f"{s.get('irr',0):.1f}%",
             f"${s.get('annualSaving',0):,.0f}"]
            for s in wf.get('scenarios', {}).values()
        ])

        export_path = os.path.join(os.path.dirname(__file__), 'data', 'export.xlsx')
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        wb.save(export_path)
        return send_file(export_path, as_attachment=True, download_name='ServiceEdge_Export.xlsx')

    except ImportError:
        return jsonify({'error':'openpyxl not installed'}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error':str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════

def _build_channel_mix(queues):
    mix = {}
    for q in queues:
        ch = q['channel']
        if ch not in mix: mix[ch] = {'channel':ch,'volume':0,'csat_w':0,'aht_w':0,'cpc_w':0}
        mix[ch]['volume'] += q['volume']
        mix[ch]['csat_w'] += q.get('csat',0) * q['volume']
        mix[ch]['aht_w'] += q.get('aht',0) * q['volume']
        mix[ch]['cpc_w'] += q.get('cpc',0) * q['volume']
    total = sum(m['volume'] for m in mix.values()) or 1
    result = []
    for m in mix.values():
        v = m['volume']
        aht_min = m['aht_w'] / max(v, 1)  # weighted average in minutes
        result.append({'channel':m['channel'],'volume':v,'pct':round(v/total*100,1),
                       'avgCSAT':round(m['csat_w']/max(v,1),2),
                       'avgAHT':round(aht_min * 60, 0),  # convert to seconds for display
                       'avgAHT_min':round(aht_min, 1),    # also keep minutes for reference
                       'avgCPC':round(m['cpc_w']/max(v,1),2)})
    result.sort(key=lambda x: x['volume'], reverse=True)
    return result

def _build_bu_mix(queues):
    mix = {}
    for q in queues:
        bu = q.get('bu','Unknown')
        if bu not in mix: mix[bu] = {'bu':bu,'volume':0,'csat_w':0}
        mix[bu]['volume'] += q['volume']
        mix[bu]['csat_w'] += q.get('csat',0) * q['volume']
    total = sum(m['volume'] for m in mix.values()) or 1
    return sorted([{'bu':m['bu'],'volume':m['volume'],'pct':round(m['volume']/total*100,1),
                    'avgCSAT':round(m['csat_w']/max(m['volume'],1),2)} for m in mix.values()],
                  key=lambda x: x['volume'], reverse=True)

def _build_intent_mix(queues):
    mix = {}
    for q in queues:
        intent = q.get('intent','Unknown')
        if intent not in mix: mix[intent] = {'intent':intent,'volume':0,'csat_w':0}
        mix[intent]['volume'] += q['volume']
        mix[intent]['csat_w'] += q.get('csat',0) * q['volume']
    total = sum(m['volume'] for m in mix.values()) or 1
    return sorted([{'intent':m['intent'],'volume':m['volume'],'pct':round(m['volume']/total*100,1),
                    'avgCSAT':round(m['csat_w']/max(m['volume'],1),2)} for m in mix.values()],
                  key=lambda x: x['volume'], reverse=True)

def _build_heatmap(queues):
    return [{'bu':q.get('bu',''),'intent':q.get('intent',''),'channel':q['channel'],
             'volume':q['volume'],'csat':q.get('csat',0),'aht':q.get('aht',0),
             'fcr':q.get('fcr',0),'cpc':q.get('cpc',0)} for q in queues]

def _build_cost_breakdown(data):
    total = data.get('totalCost',1) or 1
    return sorted([{'role':r['role'],'headcount':r['headcount'],'costPerFTE':r['costPerFTE'],
                    'totalCost':round(r['headcount']*r['costPerFTE']),
                    'pct':round(r['headcount']*r['costPerFTE']/total*100,1)} for r in data['roles']],
                  key=lambda x: x['totalCost'], reverse=True)


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
